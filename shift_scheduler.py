import ShiftScheduler # type: ignore
import time
import openpyxl # type: ignore
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side # type: ignore

# --- 設定 ---
# ※ DAYS はExcelから自動取得するので削除
POPULATION_SIZE = 50000
GENERATIONS = 1000     

INPUT_FILE = "staff_request.xlsx"
OUTPUT_FILE = "shift_result.xlsx"

# --- データ読み込み関数 (日付リストも取得するように変更) ---
def load_data_from_excel(filename):
    print(f"📂 '{filename}' からデータを読み込んでいます...")
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    roles = {}
    constraints = {}
    names = {}
    staff_count = 0
    
    # ★ここが変更点: Excelの1行目(ヘッダー)を見て日数を数える
    # 1列目=ID, 2=名前, 3=役職, 4列目からが日付
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    # 日付のカラム数 = 全カラム数 - 3 (ID,名前,役職)
    days_count = len(headers) - 3
    if days_count < 1:
        raise ValueError("Excelに日付の列がありません！テンプレートを確認してください。")
    
    date_labels = headers[3:] # 結果出力用に保存しておく("11/26(水)"などの文字)

    # 2行目からデータを読む
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        # 行が空ならスキップ
        if row[0] is None: continue

        staff_id = row[0]
        name = row[1]
        role = row[2]
        
        names[staff_id] = name
        roles[staff_id] = role
        staff_count += 1

        # 日付ごとの条件チェック
        for d in range(days_count):
            # 4列目(インデックス3)からデータが始まる
            cell_value = row[d + 3]
            
            if cell_value == "NG":
                constraints[(staff_id, d)] = "NG"
            elif cell_value == "朝":
                constraints[(staff_id, d)] = "NO_NIGHT"
            elif cell_value == "夜":
                constraints[(staff_id, d)] = "NO_MORNING"

    print(f"✅ 読み込み完了: {staff_count}人 / 期間 {days_count}日間")
    return staff_count, days_count, roles, constraints, names, date_labels

# --- 保存関数 (日付ヘッダーを反映) ---
def save_to_excel(schedule, roles_list, names_dict, date_labels, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "シフト表"
    days = len(date_labels)

    # スタイル
    fill_morning = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    fill_night = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    fill_holiday = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    fill_header = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
    # 土日のヘッダー色
    fill_sat = PatternFill(start_color="000088", end_color="000088", fill_type="solid")
    fill_sun = PatternFill(start_color="880000", end_color="880000", fill_type="solid")
    
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # ヘッダー書き込み (保存しておいた日付ラベルを使う)
    headers = ["ID", "名前", "役職"] + date_labels
    ws.append(headers)
    
    # ヘッダー装飾
    for idx, cell in enumerate(ws[1]):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        
        # 土日の色付け (文字列の中に「土」「日」が含まれているかで簡易判定)
        cell_text = str(cell.value)
        if "土" in cell_text:
            cell.fill = fill_sat
        elif "日" in cell_text:
            cell.fill = fill_sun
        else:
            cell.fill = fill_header

    shift_map = {0: "休", 1: "朝", 2: "夜"}
    
    for i, row in enumerate(schedule):
        name = names_dict.get(i, f"Staff{i}")
        excel_row = [i, name, roles_list[i]] + [shift_map[x] for x in row]
        ws.append(excel_row)

        current_row_num = i + 2
        for col_idx, val in enumerate(row):
            cell = ws.cell(row=current_row_num, column=col_idx + 4)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            
            if val == 0:
                cell.fill = fill_holiday
                cell.font = Font(color="888888")
            elif val == 1:
                cell.fill = fill_morning
            elif val == 2:
                cell.fill = fill_night
                cell.font = Font(bold=True, color="CC0000")

    # 幅調整
    ws.column_dimensions["B"].width = 15
    for col in range(4, days + 4):
        col_letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[col_letter].width = 5

    wb.save(filename)
    print(f"\n💾 Excelファイルに保存しました: {filename}")

def main():
    print(f"--- シフト生成開始 (Excel連携版) ---")
    
    # 1. ロード (days_count と date_labels も取得)
    staff_count, days_count, roles_dict, constraints, names_dict, date_labels = load_data_from_excel(INPUT_FILE)

    print(f"設定: {staff_count}人 x {days_count}日 / 個体数{POPULATION_SIZE}")

    roles_list = [roles_dict[i] for i in range(staff_count)]

    start_time = time.time()

    # 2. Rust実行 (期間 days_count を渡す)
    result_schedule, score = ShiftScheduler.run_genetic_algorithm(
        roles_list,
        constraints,
        days_count, # ここが自動で変わる
        staff_count,
        POPULATION_SIZE,
        GENERATIONS
    )

    end_time = time.time()
    print(f"処理完了！ 経過時間: {end_time - start_time:.2f}秒")
    print(f"最終スコア: {score}")

    # 3. 保存 (date_labels を渡す)
    save_to_excel(result_schedule, roles_list, names_dict, date_labels, OUTPUT_FILE)

if __name__ == "__main__":
    main()