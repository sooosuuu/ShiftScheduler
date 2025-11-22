import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import datetime
import calendar

# 日本語の曜日リスト
WEEKDAYS = ["月", "火", "水", "木", "金", "土", "日"]

# ★ここで人数と内訳を指定します！
ROLE_CONFIG = [
    ("Chief",  5),  # Chief 5人
    ("Leader", 2),  # Leader 2人
    ("Staff",  3),  # Staff 3人
    ("Assist", 10)  # Assist 10人
]

def get_shift_period(year, shift_month):
    end_date = datetime.date(year, shift_month, 25)
    if shift_month == 1:
        start_year = year - 1
        start_month = 12
    else:
        start_year = year
        start_month = shift_month - 1
    current_date = datetime.date(start_year, start_month, 26)
    
    dates = []
    while current_date <= end_date:
        dates.append(current_date)
        current_date += datetime.timedelta(days=1)
    return dates

def create_template():
    print("📅 人数指定版テンプレートを作成します")
    
    # 役職リストを展開して、スタッフ全員分のリストを作る
    # 例: ["Chief", "Chief"..., "Leader", ..., "Assist"...]
    all_staff_roles = []
    for role_name, count in ROLE_CONFIG:
        all_staff_roles.extend([role_name] * count)
    
    total_staff = len(all_staff_roles)
    print(f"設定されたスタッフ総数: {total_staff}名")

    try:
        input_year = int(input("年を入力 (例: 2025): "))
        input_month = int(input("月を入力 (例: 12): "))
    except ValueError:
        print("エラー: 数字で入力してください。")
        return

    date_list = get_shift_period(input_year, input_month)
    days_count = len(date_list)
    
    print(f"\n【{input_month}月度シフト】期間: {days_count}日間")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "希望シフト入力"

    # --- ヘッダー作成 ---
    headers = ["ID", "名前", "役職"]
    for d in date_list:
        wd_str = WEEKDAYS[d.weekday()]
        headers.append(f"{d.month}/{d.day}({wd_str})")
    ws.append(headers)

    # --- デザイン調整 ---
    fill_sat = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")
    fill_sun = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    fill_header_def = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    for col_idx, cell in enumerate(ws[1]):
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        if col_idx < 3:
            cell.fill = fill_header_def
        else:
            date_obj = date_list[col_idx - 3]
            if date_obj.weekday() == 5: cell.fill = fill_sat
            elif date_obj.weekday() == 6: cell.fill = fill_sun
            else: cell.fill = fill_header_def

    # --- ドロップダウン設定 ---
    dv = DataValidation(type="list", formula1='"NG,朝,夜"', allow_blank=True)
    dv.prompt = 'リストから選択'
    ws.add_data_validation(dv)

    # --- スタッフ行の生成 (指定された役職リスト順) ---
    for i, role in enumerate(all_staff_roles):
        # 名前もわかりやすく "Chief-0", "Assist-9" みたいにする
        # もちろんExcel上で手書き修正可能です
        name = f"{role}-{i}"
        
        row = [i, name, role] + [""] * days_count
        ws.append(row)

        # ドロップダウン適用
        current_row = i + 2
        for col in range(4, days_count + 4):
            cell = ws.cell(row=current_row, column=col)
            dv.add(cell)

    # 列幅調整
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 10
    for col in range(4, days_count + 4):
        col_letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[col_letter].width = 6

    filename = "staff_request.xlsx"
    wb.save(filename)
    print(f"\n✅ '{filename}' を作成しました！")
    print(f"内訳: Chief:5, Leader:2, Staff:3, Assist:10 (計{total_staff}名)")

if __name__ == "__main__":
    create_template()