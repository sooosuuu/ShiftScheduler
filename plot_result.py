import ShiftScheduler
import time
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# --- 設定 (実験用に少し軽めにしてもOKです) ---
STAFF_COUNT = 50       
DAYS = 30              
POPULATION_SIZE = 5000 
GENERATIONS = 1000     

# 役職設定
STAFF_ROLES = {}
for i in range(STAFF_COUNT):
    if i < STAFF_COUNT * 0.1: STAFF_ROLES[i] = "Chief"
    elif i < STAFF_COUNT * 0.3: STAFF_ROLES[i] = "Leader"
    elif i < STAFF_COUNT * 0.8: STAFF_ROLES[i] = "Staff"
    else: STAFF_ROLES[i] = "Assist"

STAFF_CONSTRAINTS = { (0, 0): "NG" }

# --- ★新機能: Excel出力関数 ---
def save_to_excel(schedule, roles_list, filename="shift_result.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "シフト表"

    # --- スタイル定義 ---
    fill_morning = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid") # 朝: 水色
    fill_night = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")   # 夜: ピンク
    fill_holiday = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid") # 休: グレー
    
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # --- ヘッダー書き込み ---
    headers = ["ID", "役職"] + [f"{d+1}日" for d in range(DAYS)]
    ws.append(headers)
    
    # ヘッダーの装飾
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # --- データ書き込み ---
    shift_map = {0: "休", 1: "朝", 2: "夜"}
    
    for i, row in enumerate(schedule):
        # 行データ作成
        excel_row = [i, roles_list[i]] + [shift_map[x] for x in row]
        ws.append(excel_row)

        # 今書き込んだ行のセルを取得して色付け
        current_row_num = i + 2 # ヘッダーが1行目なので+2
        
        for col_idx, val in enumerate(row):
            cell = ws.cell(row=current_row_num, column=col_idx + 3) # ID,役職の次から
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            
            if val == 0:
                cell.fill = fill_holiday
                cell.font = Font(color="888888")
            elif val == 1:
                cell.fill = fill_morning
                cell.font = Font(color="000000")
            elif val == 2:
                cell.fill = fill_night
                cell.font = Font(bold=True, color="CC0000")

    # 列幅調整
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 10
    for col in range(3, DAYS + 3):
        col_letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[col_letter].width = 4

    wb.save(filename)
    print(f"\n💾 Excelファイルに保存しました: {filename}")

def main():
    print(f"--- シフト生成開始 (Rust完全並列化版) ---")
    roles_list = [STAFF_ROLES[i] for i in range(STAFF_COUNT)]

    start_time = time.time()

    # Rust実行
    result_schedule, score = ShiftScheduler.run_genetic_algorithm(
        roles_list,
        STAFF_CONSTRAINTS,
        DAYS,
        STAFF_COUNT,
        POPULATION_SIZE,
        GENERATIONS
    )

    end_time = time.time()
    print(f"処理完了！ 経過時間: {end_time - start_time:.2f}秒")
    print(f"最終スコア: {score}")

    # ★Excel保存
    save_to_excel(result_schedule, roles_list)

if __name__ == "__main__":
    main()