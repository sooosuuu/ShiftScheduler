import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, simpledialog
import threading
import time
import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import unicodedata
import json
import os

import ShiftScheduler

# --- デフォルト設定 ---
DEFAULT_CONFIG = {
    "population_size": 50000,
    "generations": 1000,
    "default_roles": {
        "Chief": 5, "Leader": 2, "Staff": 3, "Assist": 10
    }
}
WEEKDAYS = ["月", "火", "水", "木", "金", "土", "日"]

class ShiftApp:
    def __init__(self, root):
        self.root = root
        self.root.title("シフト作成AI")
        self.root.geometry("700x650")

        # 設定読み込み
        self.config = self.load_config()
        
        # --- エリア1: テンプレート作成 ---
        frame_step1 = tk.LabelFrame(root, text="Step 1: 入力用ファイルの作成", padx=10, pady=10)
        frame_step1.pack(fill="x", padx=10, pady=5)
        
        lbl_desc1 = tk.Label(frame_step1, text="設定ファイル(config.json)の人数構成に基づいて、来月のひな形を作成します。")
        lbl_desc1.pack(anchor="w")
        
        btn_template = tk.Button(frame_step1, text="テンプレート作成 (staff_request.xlsx)", command=self.create_template_flow, bg="#e0f7fa")
        btn_template.pack(fill="x", pady=5)

        # --- エリア2: シフト生成 ---
        frame_step2 = tk.LabelFrame(root, text="Step 2: シフト自動生成", padx=10, pady=10)
        frame_step2.pack(fill="x", padx=10, pady=5)

        tk.Label(frame_step2, text="入力ファイル:").pack(side="left")
        self.entry_path = tk.Entry(frame_step2, width=40)
        self.entry_path.insert(0, "staff_request.xlsx")
        self.entry_path.pack(side="left", padx=5)
        
        btn_browse = tk.Button(frame_step2, text="参照...", command=self.browse_file)
        btn_browse.pack(side="left")
        
        frame_run = tk.Frame(root, pady=5)
        frame_run.pack(fill="x", padx=10)
        self.btn_run = tk.Button(frame_run, text="シフト生成開始 (Rust実行)", command=self.start_generation, bg="#ffccbc", font=("Meiryo", 12, "bold"))
        self.btn_run.pack(fill="x", ipady=5)

        # --- エリア3: ログ ---
        self.log_area = scrolledtext.ScrolledText(root, state='disabled', height=15)
        self.log_area.pack(fill="both", expand=True, padx=10, pady=10)

        self.status_var = tk.StringVar()
        self.status_var.set("待機中")
        tk.Label(root, textvariable=self.status_var, bd=1, relief=tk.SUNKEN, anchor="w").pack(side="bottom", fill="x")

    def log(self, message):
        self.log_area.config(state='normal')
        self.log_area.insert(tk.END, message + "\n")
        self.log_area.see(tk.END)
        self.log_area.config(state='disabled')

    def load_config(self):
        if os.path.exists("config.json"):
            try:
                with open("config.json", "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception as e:
                self.log(f"設定読み込みエラー: {e}")
        return DEFAULT_CONFIG

    # ---------------------------------------------------------
    #  機能1: テンプレート作成
    # ---------------------------------------------------------
    def create_template_flow(self):
        self.config = self.load_config()

        now = datetime.datetime.now()
        next_month = now.month + 1 if now.month < 12 else 1
        next_year_val = now.year + 1 if now.month == 12 else now.year
        
        target_year = simpledialog.askinteger("設定", "作成する【年】を入力してください", initialvalue=next_year_val)
        if not target_year: return
        target_month = simpledialog.askinteger("設定", "作成する【月】を入力してください", initialvalue=next_month, minvalue=1, maxvalue=12)
        if not target_month: return

        roles_dict = self.config.get("default_roles", DEFAULT_CONFIG["default_roles"])
        
        role_config = []
        total = 0
        for role in ["Chief", "Leader", "Staff", "Assist"]:
            count = roles_dict.get(role, 0)
            role_config.append((role, count))
            total += count

        if total == 0:
            messagebox.showerror("エラー", "config.json のスタッフ人数が合計0人になっています。")
            return

        try:
            self.generate_excel_template(target_year, target_month, role_config)
            messagebox.showinfo("成功", f"テンプレートを作成しました！(計{total}名)\nファイル: staff_request.xlsx")
            self.log(f"テンプレート作成完了: {target_year}年{target_month}月度 (計{total}名)")
        except Exception as e:
            messagebox.showerror("エラー", f"作成失敗: {e}")

    def generate_excel_template(self, year, shift_month, role_config):
        end_date = datetime.date(year, shift_month, 25)
        if shift_month == 1:
            start_year = year - 1
            start_month = 12
        else:
            start_year = year
            start_month = shift_month - 1
        current_date = datetime.date(start_year, start_month, 26)
        
        date_list = []
        while current_date <= end_date:
            date_list.append(current_date)
            current_date += datetime.timedelta(days=1)
        days_count = len(date_list)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "希望シフト入力"

        headers = ["ID", "名前", "役職"]
        for d in date_list:
            headers.append(f"{d.month}/{d.day}({WEEKDAYS[d.weekday()]})")
        ws.append(headers)

        fill_sat = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")
        fill_sun = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        fill_header = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col_idx, cell in enumerate(ws[1]):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            if col_idx < 3: cell.fill = fill_header
            else:
                d = date_list[col_idx - 3]
                if d.weekday() == 5: cell.fill = fill_sat
                elif d.weekday() == 6: cell.fill = fill_sun
                else: cell.fill = fill_header

        dv = DataValidation(type="list", formula1='"NG,朝,夜"', allow_blank=True)
        ws.add_data_validation(dv)

        all_staff_roles = []
        for role_name, count in role_config:
            all_staff_roles.extend([role_name] * count)

        for i, role in enumerate(all_staff_roles):
            name = f"{role}-{i}"
            ws.append([i, name, role] + [""] * days_count)
            current_row = i + 2
            for col in range(4, days_count + 4):
                cell = ws.cell(row=current_row, column=col)
                dv.add(cell)

        ws.column_dimensions["B"].width = 15
        for col in range(4, days_count + 4):
            col_letter = openpyxl.utils.get_column_letter(col)
            ws.column_dimensions[col_letter].width = 6

        wb.save("staff_request.xlsx")

    # ---------------------------------------------------------
    #  機能2: シフト生成
    # ---------------------------------------------------------
    def browse_file(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filename:
            self.entry_path.delete(0, tk.END)
            self.entry_path.insert(0, filename)

    def start_generation(self):
        input_file = self.entry_path.get()
        if not os.path.exists(input_file):
            messagebox.showerror("エラー", "入力ファイルが見つかりません")
            return

        self.btn_run.config(state="disabled", text="計算中... (Rust稼働中)")
        self.status_var.set("計算中...")
        self.log("--- シフト生成プロセスを開始 ---")
        threading.Thread(target=self.run_logic, args=(input_file,)).start()

    def run_logic(self, input_file):
        try:
            self.log("データを読み込んでいます...")
            staff_count, days_count, roles, constraints, names, date_labels = self.load_data_clean(input_file)
            self.log(f"読み込み完了: {staff_count}名 / {days_count}日間")

            roles_list = [roles[i] for i in range(staff_count)]
            
            pop_size = self.config.get("population_size", 50000)
            gens = self.config.get("generations", 1000)

            self.log(f"Rustエンジン起動 (個体数:{pop_size})...")
            start_time = time.time()
            
            result_schedule, score = ShiftScheduler.run_genetic_algorithm(
                roles_list, constraints, days_count, staff_count, pop_size, gens
            )
            
            elapsed = time.time() - start_time
            self.log(f"計算完了: {elapsed:.2f}秒 (スコア: {score})")

            self.analyze_and_report(result_schedule, roles_list, constraints, days_count, staff_count, names, date_labels)

            # ★変更点: 日時付きのファイル名を生成
            now_str = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = f"shift_result_{now_str}.xlsx"
            
            self.save_data(result_schedule, roles_list, names, date_labels, output_file)
            self.log(f"保存完了: {output_file}")
            
            messagebox.showinfo("完了", f"シフト作成が完了しました！\n\n保存ファイル:\n{output_file}")
            os.startfile(output_file)

        except Exception as e:
            self.log(f"エラー: {e}")
            messagebox.showerror("エラー", str(e))
        finally:
            self.root.after(0, self.reset_gui)

    def reset_gui(self):
        self.btn_run.config(state="normal", text="シフト生成開始 (Rust実行)")
        self.status_var.set("待機中")

    def load_data_clean(self, filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        days_count = len(headers) - 3
        date_labels = headers[3:]
        roles = {}
        constraints = {}
        names = {}
        staff_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None: continue
            staff_id, name, role = row[0], row[1], row[2]
            names[staff_id] = name
            roles[staff_id] = role
            staff_count += 1
            for d in range(days_count):
                val = row[d+3]
                if val:
                    val = unicodedata.normalize('NFKC', str(val)).strip()
                    if val in ["NG", "ng", "休み", "×"]: constraints[(staff_id, d)] = "NG"
                    elif val in ["朝", "Morning", "早番"]: constraints[(staff_id, d)] = "NO_NIGHT"
                    elif val in ["夜", "Night", "遅番"]: constraints[(staff_id, d)] = "NO_MORNING"
        return staff_count, days_count, roles, constraints, names, date_labels

    def analyze_and_report(self, schedule, roles, constraints, days, staff_count, names, date_labels):
        self.log("\n--- シフト診断レポート ---")
        issues = 0
        for (sid, d), c_type in constraints.items():
            shift = schedule[sid][d]
            if (c_type == "NG" and shift != 0) or (c_type == "NO_MORNING" and shift == 1) or (c_type == "NO_NIGHT" and shift == 2):
                self.log(f"❌ [希望違反] {names.get(sid)} {date_labels[d]}")
                issues += 1
        
        for sid, row in enumerate(schedule):
            consecutive = 0
            for d, s in enumerate(row):
                if s != 0:
                    consecutive += 1
                    if consecutive > 6:
                         self.log(f"⚠️ [過労] {names.get(sid)}: {consecutive}連勤")
                         issues += 1
                else: consecutive = 0
            for d in range(days):
                if row[d] == 2:
                    if d+1 < days and row[d+1] == 1:
                        self.log(f"⚠️ [休息不足] {names.get(sid)}: 夜→朝")
                        issues += 1
                    elif d+2 < days and row[d+2] == 1:
                        self.log(f"⚠️ [休息不足] {names.get(sid)}: 夜→休→朝(間隔不足)")
                        issues += 1

        for d in range(days):
            morning = sum(1 for i in range(staff_count) if schedule[i][d] == 1)
            night = sum(1 for i in range(staff_count) if schedule[i][d] == 2)
            if morning < 5:
                self.log(f"⚠️ [人手不足] {date_labels[d]}: 朝{morning}人")
                issues += 1
            if night < 5:
                self.log(f"⚠️ [人手不足] {date_labels[d]}: 夜{night}人")
                issues += 1
        
        if issues == 0: self.log("✨ 違反箇所ゼロ！完璧です。")
        else: self.log(f"計 {issues} 件の課題あり")
        self.log("------------------------")

    def save_data(self, schedule, roles_list, names, date_labels, filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "シフト表"
        fill_morning = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
        fill_night = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        fill_holiday = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        fill_sat = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")
        fill_sun = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        ws.append(["ID", "名前", "役職"] + date_labels)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
            txt = str(cell.value)
            if "土" in txt: cell.fill = fill_sat
            elif "日" in txt: cell.fill = fill_sun
            else: cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        shift_map = {0: "休", 1: "朝", 2: "夜"}
        for i, row in enumerate(schedule):
            name = names.get(i, f"Staff{i}")
            ws.append([i, name, roles_list[i]] + [shift_map[x] for x in row])
            for col_idx, val in enumerate(row):
                cell = ws.cell(row=i+2, column=col_idx+4)
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
                if val == 0: 
                    cell.fill = fill_holiday
                    cell.font = Font(color="888888")
                elif val == 1: cell.fill = fill_morning
                elif val == 2: 
                    cell.fill = fill_night
                    cell.font = Font(bold=True, color="CC0000")
        wb.save(filename)

if __name__ == "__main__":
    root = tk.Tk()
    app = ShiftApp(root)
    root.mainloop()